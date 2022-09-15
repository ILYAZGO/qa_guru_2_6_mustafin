import zipfile,PyPDF2
from openpyxl import load_workbook
from .packing import packing

def test_from_zip():
    packing()
    assert zipfile.is_zipfile('all.zip') == True

def test_csv():
    with zipfile.ZipFile('all.zip') as all:
        with all.open('csv_example.csv', mode='r') as file:
            csv_data=file.read().decode()
    assert csv_data.find('Иванов') != -1
    assert csv_data.find('Кадыров') != 1

def test_pdf():
    with zipfile.ZipFile('all.zip') as all:
        with all.open('book.pdf', mode='r') as file:
            pdfReader = PyPDF2.PdfFileReader(file)
            pages_count=pdfReader.numPages
            text=''
            for i in range(0, pages_count):
                pdfcount= pdfReader.getPage(i)
                text += pdfcount.extractText() + "\n"
    assert text.find('Python') != -1
    assert text.find('Clojure') != 1


def test_xlsx():
    with zipfile.ZipFile('all.zip') as all:
        with all.open('База данных Жители Кременчуга и Кременчугского уезда.xlsx', mode='r') as file:
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            text = ''
            for row in ws.iter_rows(min_row=ws.min_row,max_row=ws.max_row):
                for cell in row:
                    text += str(cell.value)+" "
    assert text.find('Ячмяник') != -1
    assert text.find('Кадыров') != 1